import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

def getCurrentTime():
    now = datetime.now()
    return now.strftime("%H:%M:%S")

menuHarga = {
    "Nasi Goreng": 15000,
    "Mie Goreng": 12000,
    "Ayam Bakar": 20000,
    "Sate Ayam": 18000,
    "Bakso": 13000,
}

def updateTotal():
    try:
        totalPembayaran = 0
        for makanan, harga in menuHarga.items():
            jumlah = int(entriesJumlah[makanan].get())
            totalPembayaran += harga * jumlah
        entryTotal.config(state=tk.NORMAL)
        entryTotal.delete(0, tk.END)
        entryTotal.insert(0, str(totalPembayaran))
        entryTotal.config(state=tk.DISABLED)
    except ValueError:
        entryTotal.config(state=tk.NORMAL)
        entryTotal.delete(0, tk.END)
        entryTotal.config(state=tk.DISABLED)

def saveToExcel():
    jam = getCurrentTime()
    filePath = 'laporan_penjualan.xlsx'
    if os.path.exists(filePath):
        wb = load_workbook(filePath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Makanan', 'Jumlah', 'Total Biaya', 'Jam'])
    
    for makanan, jumlahEntry in entriesJumlah.items():
        jumlah = int(jumlahEntry.get())
        if jumlah > 0:
            harga = menuHarga[makanan]
            totalBiaya = harga * jumlah
            ws.append([makanan, jumlah, totalBiaya, jam])
    
    wb.save(filePath)
    messagebox.showinfo("Sukses", "Data berhasil disimpan!")

root = tk.Tk()
root.title("Aplikasi Laporan Penjualan")
root.geometry("800x600")
width = root.winfo_screenwidth()
height = root.winfo_screenheight()
x = (width // 2) - (800 // 2)
y = (height // 2) - (600 // 2)
root.geometry(f"800x600+{x}+{y}")

root.configure(bg='#f0f0f0')

# Jam di pojok kanan atas
jamFrame = tk.Frame(root, bg='#f0f0f0')
jamFrame.pack(fill='x', pady=10)
tk.Label(jamFrame, text="Jam:", font=("Poppins", 14), bg='#f0f0f0').pack(side='right', padx=5)
entryJam = tk.Label(jamFrame, text=getCurrentTime(), font=("Poppins", 14), bg='#f0f0f0')
entryJam.pack(side='right', padx=5)

container = tk.Frame(root, bg='#f0f0f0')
container.pack(expand=True, padx=20, pady=20, fill='both')

entriesJumlah = {}

# Frame untuk menampilkan card makanan
foodContainer = tk.Frame(container, bg='#f0f0f0')
foodContainer.pack(pady=20)

def createFoodCard(menu, price):
    frame = tk.Frame(foodContainer, bg='white', bd=2, relief='solid')
    frame.pack(side='left', padx=10, pady=10)
    frame.configure(highlightbackground="black", highlightcolor="black", highlightthickness=1)
    
    label = tk.Label(frame, text=menu, font=("Poppins", 12))
    label.pack(side='top', pady=5)
    
    labelPrice = tk.Label(frame, text=f"Rp {price}", font=("Poppins", 10))
    labelPrice.pack(side='top', pady=5)
    
    jumlahEntry = tk.Entry(frame, width=5)
    jumlahEntry.pack(side='top', pady=5)
    jumlahEntry.insert(0, "0")
    jumlahEntry.bind("<KeyRelease>", lambda e: updateTotal())
    entriesJumlah[menu] = jumlahEntry

for makanan, harga in menuHarga.items():
    createFoodCard(makanan, harga)

# Kolom untuk total pembayaran
totalFrame = tk.Frame(container, bg='#f0f0f0')
totalFrame.pack(pady=20)

tk.Label(totalFrame, text="Total Pembayaran:", font=("Poppins", 14), bg='#f0f0f0').pack(side='left', padx=20, pady=10)
entryTotal = tk.Entry(totalFrame, font=("Poppins", 14), bg='#ffffff', relief='solid', state=tk.DISABLED)
entryTotal.pack(side='left', padx=20, pady=10)

# Tombol untuk menyimpan pesanan
saveButton = tk.Button(container, text="Simpan Pesanan", font=("Poppins", 14), bg='#4caf50', fg='#ffffff', command=saveToExcel, relief='raised', bd=2)
saveButton.pack(pady=10)

def updateTime():
    currentTime = getCurrentTime()
    entryJam.config(text=currentTime)
    root.after(1000, updateTime)

updateTime()

root.mainloop()
